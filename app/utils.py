# -*- coding: utf-8 -*-
from datetime import datetime
from app.models import Logs, session1, Lots, Cost_center
from functools import wraps
from flask import session, redirect
from app.models import IP_HOME
from config import main_dir_docs, main_dir
# from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
import smtplib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# Authentication
def requires_auth(f):
    @wraps(f)
    def decorated_function(*args):
        if session['rol'] == 'None' or session['rol'] is None or session['rol'] == '' or session['rol'] == 'Usuario no encontrado':
            url = f'{IP_HOME}logout/No tens permisos per entrar en aquesta a plicació'
            return redirect(url)
        elif session['acronim'] == 'None' or session['acronim'] is None or session['acronim'] == '' or session['acronim'] == 'Usuario no encontrado':
            url = f"{IP_HOME}logout/L'arcronim s'ha esborrat, siusplau torna a fer el log-in"
            return redirect(url)
        else:
            pass
        return f(*args)
    return decorated_function


def instant_date():
    '''
    1 - Obtenim la data actual
    2 - La formatem perquè estigui com nosaltres volem
    3 - Enviem la data formatada.

    :return: La data actual formatada.
    :rtype: string

    '''
    date = datetime.now()
    format_date = date.strftime("%d-%m-%Y")
    return format_date


def year_now():
    '''
        Agafem l'any actual i el retornem en YYYY

        :return: Any actual en el format YYYY
        :rtype: string
    '''
    year = datetime.now().year
    return year


def save_log(dict_info_lot):
    '''
        1 - Afegim a la BD de Logs una linia amb la info que ens donen

        :param str dict_info_lot: Diccionari amb la informació requerida

        :return: True o False.
        :rtype: json

    '''
    try:
        insert_log = Logs(id_lot=dict_info_lot['id_lot'],
                          type=dict_info_lot['type'],
                          info=dict_info_lot['info'],
                          user=dict_info_lot['user'],
                          id_user=dict_info_lot['id_user'],
                          date=dict_info_lot['date'])
        session1.add(insert_log)
        # session1.commit()
    except Exception:
        return "False"
    return 'True'


def list_desciption_lots():
    '''
        1 - Agafem tots el lots actius i els retornem.

        :return: llista amb els objectes de tipo Lots que hem trobat a la BD
        :rtype: llista d'objectes
    '''
    select_lot = session1.query(Lots).filter(Lots.active == 1).all()
    return select_lot


def list_cost_center():
    '''
        1 - Agafem tots els camps de cost center

        :return: llista amb els objectes de tipo Cost center que hem trobat a la BD
        :rtype: llista d'objectes
    '''
    select_cost_center = session1.query(Cost_center).all()
    return select_cost_center


def create_excel(select_row):
    '''
        Creem un csv amb la informació que recollim de la BD
        Iterem la llista i anem posant la informació on toca
    '''
    try:
        # # Crear arxiu nou
        # archivo = f"{main_dir_docs}/comandes_pendents.csv"
        # csv = open(archivo, "w")
        # # Inserir linies al csv
        # csv.write('Peticionari;Codi proveidor;Descripció;Codi SAP;Codi LOG;Unitats;Data creació;Usuari;CECO\n')
        # for command, lot in select_row:
        #     linia_csv = str(command.user_create) + ';'
        #     linia_csv += str(lot.catalog_reference) + ';'
        #     linia_csv += str(lot.description) + ';'
        #     linia_csv += str(lot.code_SAP) + ';'
        #     linia_csv += str(lot.code_LOG) + ';'
        #     linia_csv += str(command.units) + ';'
        #     linia_csv += str(command.date_create) + ';'
        #     linia_csv += str(command.user_create) + ';'
        #     linia_csv += str(command.cost_center) + ';'
        #     linia_csv += '\n'
        #     csv.write(linia_csv)
        # csv.close()

        def create_dataframe(select_row):
            # Crear un DataFrame con los datos
            data = {
                'Peticionari': [],
                'Codi proveidor': [],
                'Descripció': [],
                'Codi SAP': [],
                'Codi LOG': [],
                'Unitats': [],
                'Data creació': [],
                'Usuari': [],
                'CECO': [],
            }

            for command, lot in select_row:
                data['Peticionari'].append(command.user_create)
                data['Codi proveidor'].append(lot.catalog_reference)
                data['Descripció'].append(lot.description)
                data['Codi SAP'].append(lot.code_SAP)
                data['Codi LOG'].append(lot.code_LOG)
                data['Unitats'].append(command.units)
                data['Data creació'].append(command.date_create)
                data['Usuari'].append(command.user_create)
                data['CECO'].append(command.cost_center)

            return pd.DataFrame(data)

        # Crear DataFrames per l'any actual
        df_current_year = create_dataframe(select_row)

        # Guardar el DataFrame en un archivo Excel
        path = f"{main_dir_docs}/comandes_pendents.xlsx"

        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df_current_year.to_excel(writer, sheet_name="Comandes_pendents", index=False)

        ##############################################################################
        # ########################## ESTILS FULLS ####################################
        ##############################################################################

        # Ajustar automáticamente el ancho de las columnas
        wb = load_workbook(path)
        ws = wb.active

        # --- Aplicar estilos al encabezado ---
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo amarillo
        header_font = Font(size=13, bold=True)  # Letra tamaño 13 y en negrita
        for cell in ws[1]:  # Primera fila (encabezado)
            cell.fill = header_fill
            cell.font = header_font

        # --- Ajustar el tamaño de las columnas automáticamente ---
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho

        ws.column_dimensions['I'].width = 7  # Ajustar ancho
        ws.column_dimensions['D'].width = 11  # Ajustar ancho
        ws.column_dimensions['E'].width = 11  # Ajustar ancho

        # --- Ajustar altura de todas las filas (margen superior e inferior) ---
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 19  # Ajustar altura de fila
            for cell in row:
                cell.alignment = Alignment(vertical="center", horizontal="left")  # Alineación vertical centrada

        wb.save(path)  # Guardar cambios
        return True
    except Exception:
        return False


def send_mail(list_info_excel):
    '''
        1 - Cridem a create excel info reception pero que ens crei un excel
        2 - Preparem totes les dades del correu
        3 - Adjuntem l'archiu al correu
        3 - Enviem el correu amb totes les dades requerides.

        :param list list_info_excel: llista de diccionaris amb la informació requerida

        :return: None
        :rtype: None
    '''
    try:
        create_excel_info_reception(list_info_excel)

        subject = 'UDMMP | Recepció de productes'
        email_sender = "udmmp.girona.ics@gencat.cat"
        image_path = f'{main_dir}/logo.png'

        # Crear el mensaje multipart
        msg = MIMEMultipart()
        msg['From'] = email_sender
        msg['Subject'] = subject

        if list_info_excel[0]['analytical_technique'] == 'NGS':
            emails = ['monicacoll.girona.ics@gencat.cat', 'llopez@gencardio.com', 'mcorona.girona.ics@gencat.cat', 'msoriano@idibgi.org', 'mpinsach.girona.ics@gencat.cat', 'asimon.girona.ics@gencat.cat', 'nneto.girona.ics@gencat.cat', 'ferran.pico@gencardio.com']
            msg["To"] = ', '.join(emails)
        elif list_info_excel[0]['analytical_technique'] == 'Genotipat':
            emails = ['nneto.girona.ics@gencat.cat', 'mpuigmule.girona.ics@gencat.cat']
            msg["To"] = ', '.join(emails)
        elif list_info_excel[0]['analytical_technique'] == 'Sanger':
            emails = ['ferran.pico@gencardio.com', 'aardila@idibgi.org', 'nneto.girona.ics@gencat.cat', 'abatchelli.girona.ics@gencat.cat']
            msg["To"] = ', '.join(emails)
        elif list_info_excel[0]['analytical_technique'] == 'Extracció':
            emails = ['abatchelli.girona.ics@gencat.cat', 'mcorona.girona.ics@gencat.cat', 'nneto.girona.ics@gencat.cat', 'ferran.pico@gencardio.com']
            msg["To"] = ', '.join(emails)
        elif list_info_excel[0]['analytical_technique'] == 'qPCR':
            emails = ['mpuigmule.girona.ics@gencat.cat', 'nneto.girona.ics@gencat.cat', 'ferran.pico@gencardio.com', 'abatchelli.girona.ics@gencat.cat']
            msg["To"] = ', '.join(emails)
        else:
            # emails = ['asimon.girona.ics@gencat.cat', 'asimon@gencardio.com']
            # msg["To"] = ', '.join(emails)
            return

        # Adjuntar el contenido del mensaje en formato HTML
        html = f"""
            <html>
                <body>
                    <p>=-=-=- No respongueu a aquest missatge, és un correu només d'informació =-=-=-=</p>

                    <p>Benvolgut/da,</p>

                    <p>L'informem que s'han rebut els productes associats a la teva tècnica analítica. T'adjuntem un excel amb la informació.:</p>

                    <p><i>Per a qualsevol dubte, podeu contactar-nos a udmmp.tic.girona.ics@gencat.cat, asimon.girona.ics@gencat.cat, aperezp.girona.ics@gencat.cat.</i></p>

                    <p>Moltes gràcies.</p>

                    <div style="text-align: left;">
                        <img src="cid:image1" style="width:250px; height:auto; display:block; margin:0;">
                    </div>

                    <p>--</p>
                    <p>UDMMP | Unitat de Diagnóstic Molecular i Medicina Personalitzada<br>
                    Institut Català de la Salut | Generalitat de Catalunya<br>
                    Hospital Santa Caterina. Parc Hospitalari Martí i Julià<br>
                    C/Dr. Castany s/n | 17190 Salt | Tel. 972189023 | Ext. 9929</p>
                </body>
            </html>
        """
        msg.attach(MIMEText(html, 'html'))

        # Adjuntar la imagen
        with open(image_path, 'rb') as img:
            mime_image = MIMEImage(img.read())
            mime_image.add_header('Content-ID', '<image1>')
            mime_image.add_header('Content-Disposition', 'inline', filename=image_path)
            msg.attach(mime_image)

        # Adjuntar el archivo
        with open(f"{main_dir_docs}/recepcio_stock.csv", 'rb') as file:
            mime_base = MIMEBase('application', 'octet-stream')
            mime_base.set_payload(file.read())
            encoders.encode_base64(mime_base)
            mime_base.add_header('Content-Disposition', f'attachment; filename=recepcio_stock.csv')
            msg.attach(mime_base)

        # message = "S'han rebut productes associats a la teva tècnica analítica. T'adjunto un excel amb la informació."

        # subject = 'UDMMP | Recepció de productes'

        # em = EmailMessage()
        # email_sender = "udmmp.girona.ics@gencat.cat"
        # em["From"] = email_sender

        # if list_info_excel[0]['analytical_technique'] == 'NGS':
        #     emails = ['monicacoll.girona.ics@gencat.cat', 'llopez@gencardio.com', 'mcorona.girona.ics@gencat.cat', 'mmoliner@idibgi.org', 'msoriano@idibgi.org', 'mpinsach.girona.ics@gencat.cat', 'aperezs.girona.ics@gencat.cat', 'asimon.girona.ics@gencat.cat']
        #     em["To"] = ', '.join(emails)
        # elif list_info_excel[0]['analytical_technique'] == 'Genotipat2':
        #     emails = ['nneto.girona.ics@gencat.cat', 'mpuigmule.girona.ics@gencat.cat', 'mmoliner@idibgi.org']
        #     em["To"] = ', '.join(emails)
        # elif list_info_excel[0]['analytical_technique'] == 'Sanger2':
        #     emails = ['ferran.pico@gencardio.com', 'aardila@idibgi.org', 'aperezs.girona.ics@gencat.cat']
        #     em["To"] = ', '.join(emails)
        # elif list_info_excel[0]['analytical_technique'] == 'Extracció2':
        #     emails = ['abatchelli.girona.ics@gencat.cat', 'igomez.girona.ics@gencat.cat']
        #     em["To"] = ', '.join(emails)
        # else:
        #     # emails = ['asimon.girona.ics@gencat.cat', 'asimon@gencardio.com']
        #     # em["To"] = ', '.join(emails)
        #     return

        # em["Subject"] = subject
        # em.set_content(message)

        # with open(f"{main_dir_docs}/recepcio_stock.csv", 'rb') as file:
        #     file_data = file.read()
        #     file_name = 'recepcio_stock.csv'
        # em.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)

        with smtplib.SMTP("172.16.2.137", 25) as smtp:
            # smtp.sendmail(email_sender, emails, msg.as_string())
            smtp.send_message(msg)
    except Exception:
        return
    return


def create_excel_info_reception(list_info_excel):
    '''
        Creem un excel i l'omplim amb la informació de la llista que ens pasen per parametre

        :param list list_info_excel: llista de diccionaris amb la informació requerida

        :return: False o True
        :rtype: bool
    '''
    try:
        # Crear arxiu nou
        archivo = f"{main_dir_docs}/recepcio_stock.csv"
        csv = open(archivo, "w")
        # Inserir linies al csv

        if str(list_info_excel[0]['description_subreference']) == '' or str(list_info_excel[0]['description_subreference']) == 'None':
            csv.write('Usuari petició;Nom producte;Referencia producte;Lot;Lot intern;Data recepcio;Data caducitat;\n')
        else:
            csv.write('Usuari petició;Nom producte;Referencia producte;Nom producte subreferencia;Identificador subreferencia;Lot;Lot intern;Data recepcio;Data caducitat;\n')

        for info_excel in list_info_excel:
            linia_csv = str(info_excel['user_add_command']) + ';'
            linia_csv += str(info_excel['catalog_reference']) + ';'
            linia_csv += str(info_excel['description']) + ';'
            if str(info_excel['description_subreference']) == '' or str(info_excel['description_subreference']) == 'None':
                pass
            else:
                linia_csv += str(info_excel['description_subreference']) + ';'
                linia_csv += str(info_excel['id_reactive']) + ';'
            linia_csv += str(info_excel['lot']) + ';'
            linia_csv += str(info_excel['internal_lot_value']) + ';'
            linia_csv += str(info_excel['reception_date']) + ';'
            linia_csv += str(info_excel['date_expiry']) + ';'
            linia_csv += '\n'
            csv.write(linia_csv)
        csv.close()
        return True
    except Exception:
        return False


def to_dict(obj):
    """Convertir un objeto SQLAlchemy a un diccionario."""
    return {c.name: getattr(obj, c.name) for c in obj.__table__.columns}


def format_price(value):
    '''
        Aquesta funció format_price s'encarrega de normalitzar els valors numèrics per tal que tinguin un format coherent, especialment per a preus. El seu funcionament és el següent:
        Gestió de guions:
        Si el valor introduït és un guió ('-'), el retorna tal qual sense cap modificació.

        1 - Conversió a número:
            Converteix el valor a un número decimal, substituint la coma per un punt si cal. Això permet gestionar formats europeus on la coma s'utilitza com a separador decimal.

        2 - Format segons el tipus de número:
            Si el número és enter (sense decimals), es retorna com a enter.
            Si té decimals, es retorna amb màxim dos decimals.

        3 - Gestió d'errors:
            Si el valor no es pot convertir a número (per exemple, si és text), retorna el valor original.

        :param str value: valor extret de l'excel

        :return: Retorna el valor formatejat o en el seu defecte el valor original
        :rtype: int, flot o str
    '''
    # Si és un guió, retornem tal qual
    if str(value).strip() == '-':
        return value

    # Convertim a float i fem servir punt decimal
    try:
        num = float(str(value).replace(',', '.'))

        # Si és un número rodó, retornem sense decimals
        if num.is_integer():
            return int(num)
        else:
            return round(num, 2)
    except ValueError:
        # En cas d'error de format, retornem el valor original
        return value
