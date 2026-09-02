from flask import render_template, request, redirect, session, flash, jsonify, send_file, after_this_request
from app import app
from app.utils import requires_auth, list_desciption_lots, list_cost_center, to_dict, save_log, instant_date, send_mail_generic
from app.models import IP_HOME, session1, Lots, Stock_lots, Lot_consumptions, Buy_primers
import jwt
import json
import re
import urllib.error
import urllib.request
from sqlalchemy import and_, or_, outerjoin, func
from datetime import datetime
from config import main_dir, ip_address
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import main_dir_docs
import os
import tempfile
import zipfile


def parse_primer_gene_exon(sequence_name):
    match = re.match(r'^\s*([^_]+)_([0-9]{1,2}\s*-\s*[0-9]{1,2}|[0-9]{1,2}|[XY])', sequence_name or '', re.IGNORECASE)
    if not match:
        raise ValueError(f"No s'ha pogut obtenir el gen i l'exó del nom del primer: {sequence_name}")
    return match.group(1).strip(), match.group(2).strip().upper()


def parse_primer_pair_info(sequence_name):
    match = re.match(
        r'^\s*([^_]+)_([0-9]{1,2}\s*-\s*[0-9]{1,2}|[0-9]{1,2}|[XY])(as|s)(?=$|[^A-Za-z])',
        sequence_name or '',
        re.IGNORECASE
    )
    if not match:
        return None

    return {
        'gen': match.group(1).strip(),
        'exon': match.group(2).strip().upper(),
        'role': match.group(3).lower()
    }


def get_selected_primer_sense_pair(selected_primers):
    if len(selected_primers) != 2:
        return None

    parsed_primers = []
    for primer in selected_primers:
        pair_info = parse_primer_pair_info(primer.sequence_name)
        if not pair_info:
            return None
        parsed_primers.append((primer, pair_info))

    roles = {pair_info['role'] for _, pair_info in parsed_primers}
    same_gen_exon = (
        parsed_primers[0][1]['gen'].lower() == parsed_primers[1][1]['gen'].lower()
        and parsed_primers[0][1]['exon'].replace(' ', '') == parsed_primers[1][1]['exon'].replace(' ', '')
    )

    if roles != {'s', 'as'} or not same_gen_exon:
        return None

    sense = next(primer for primer, pair_info in parsed_primers if pair_info['role'] == 's')
    antisense = next(primer for primer, pair_info in parsed_primers if pair_info['role'] == 'as')
    pair_technique = f"{sense.technique or ''} {antisense.technique or ''}"

    return {
        'gen': parsed_primers[0][1]['gen'],
        'exon': parsed_primers[0][1]['exon'],
        'sequence_name': sense.sequence_name,
        'sense_name': sense.sequence_name,
        'antisense_name': antisense.sequence_name,
        'sense': sense.sequence,
        'antisense': antisense.sequence,
        'qpcr': '1' if re.search(r'qpcr', pair_technique, re.IGNORECASE) else '0',
        'm13': '1' if re.search(r'm13', f"{sense.sequence_name} {antisense.sequence_name}", re.IGNORECASE) else '0'
    }


def get_isoforma_by_gene_exon(sequence_name):
    gen, exon = parse_primer_gene_exon(sequence_name)
    url = f'{ip_address}:5016/api/isoforma_by_gen_exon'
    payload = json.dumps({'gen': gen, 'exon': exon}).encode('utf-8')
    request_api = urllib.request.Request(
        url,
        data=payload,
        headers={'Content-Type': 'application/json'},
        method='POST'
    )

    try:
        with urllib.request.urlopen(request_api, timeout=5) as response:
            response_data = json.loads(response.read().decode('utf-8') or '{}')
    except urllib.error.HTTPError as e:
        try:
            response_data = json.loads(e.read().decode('utf-8') or '{}')
            message = response_data.get('message') or response_data.get('error') or str(e)
        except Exception:
            message = str(e)
        raise ValueError(f"No s'ha pogut trobar la isoforma per {sequence_name} ({gen}_{exon}). {message}")
    except urllib.error.URLError as e:
        raise ValueError(f"No s'ha pogut connectar amb registres per consultar la isoforma de {sequence_name} ({gen}_{exon}). {e.reason}")
    except Exception as e:
        raise ValueError(f"Error consultant la isoforma de {sequence_name} ({gen}_{exon}). {e}")

    isoforma = response_data.get('isoforma')
    if not isoforma:
        raise ValueError(f"La resposta de registres no conté cap isoforma per {sequence_name} ({gen}_{exon}).")
    return isoforma


def receive_primers_by_ids(list_id_primer, received_by):
    date_now = instant_date()
    text_email = '<p style="margin-bottom:10px;">Hem rebut els seguents primers :</p>'
    text_header_email = 'Recepcio de primers'
    not_found = ''
    id_primers_change = ''
    send_mail = []

    for primer_id in list_id_primer:
        select_primer = session1.query(Buy_primers).filter(Buy_primers.id == primer_id).first()
        if select_primer is None:
            not_found += str(primer_id) + ';'
            continue

        if select_primer.buy == 1 and select_primer.received == 0 and select_primer.delete == 0:
            select_primer.received = 1
            select_primer.received_date = date_now
            select_primer.received_by = received_by

            text_email += f"&nbsp;&nbsp;&nbsp;• {select_primer.sequence_name} - {select_primer.sequence}<br>"
            id_primers_change += str(primer_id) + ';'
            if select_primer.email_send not in send_mail:
                send_mail.append(select_primer.email_send)

    return {
        'id_primers_change': id_primers_change.rstrip(';'),
        'not_found': not_found.rstrip(';'),
        'send_mail': send_mail,
        'text_email': text_email,
        'text_header_email': text_header_email
    }


# Pagina incial i visualització
@app.route('/')
@requires_auth
def main():
    '''
        Redirigeix al home de lots
    '''
    return render_template('home.html', list_desciption_lots=list_desciption_lots(),
                           list_cost_center=list_cost_center())


@app.route('/logout')
def logout():
    '''
        Redirigeix a l'applicació home/logout
    '''
    url = IP_HOME + 'logout'

    return redirect(url)


@app.route('/apps')
@requires_auth
def apps():
    '''
        Guardem les cookies en un tocken i les enviem a home/apps, perque puguin obrir cualsevol applicació a la que
        tinguin acceès.
    '''
    tocken_cookies = {'user_tok': session['user'], 'rols_tok': session['rols'], 'email_tok': session['email'],
                      'id_client_tok': session['idClient'], 'rol_tok': 'None', 'acronim_tok': session['acronim']}
    secret_key = '12345'
    token = jwt.encode(tocken_cookies, secret_key, algorithm='HS256')
    url = f'{IP_HOME}apps/token?token={token}'

    return redirect(url)


@app.route('/receive_token')
def receive_token():
    '''
        Rebem el tocken i assignem a la nostre sessions els valors.
    '''
    received_token = request.args.get('token')
    secret_key = '12345'  # Debe ser la misma clave utilizada para generar el token

    try:
        decoded_token = jwt.decode(received_token, secret_key, algorithms=['HS256'])
        session['user'] = decoded_token.get('user_tok', 'Usuario no encontrado')
        session['rols'] = decoded_token.get('rols_tok', 'Usuario no encontrado')
        session['email'] = decoded_token.get('email_tok', 'Usuario no encontrado')
        session['idClient'] = decoded_token.get('id_client_tok', 'Usuario no encontrado')
        session['rol'] = decoded_token.get('rol_tok', 'Usuario no encontrado')
        session['acronim'] = decoded_token.get('acronim_tok', 'Usuario no encontrado')
        print(session['user'])
        print(session['rols'])
        print(session['email'])
        print(session['idClient'])
        print(session['rol'])
        print(session['acronim'])
        return redirect('/')
    except Exception:
        return redirect('/logout')


@app.route('/search_lot_db', methods=['POST'])
@requires_auth
def search_lot_db():
    '''
        1 - Recollim la informació de l'ajax
        2 - Comprovem si aquest lot té història.
        2.1 - Si no en té retornem False més un missatge d'explicació per l'usuari.
        2.2 - Si és que si agafem la informació que hem trobat la posem en una llista de diccionaris.
        3 - Convertim la llista de diccionaris en un json
        4 - Retornem un True més la llista de diccionaris convertida a json.

        :param str id_lot: Identificador unit del lot

        :return: True i la llista de diccionaris amb la info o False i una explicació per l'usuari
        :rtype: json
    '''
    try:
        select_lots = session1.query(Lots).all()
        if not select_lots:
            return 'False_//_No hi ha cap lot a la BD.'

        list_lots = []
        for lot in select_lots:
            dict_lots = {}
            dict_lots['id'] = lot.key
            # dict_lots['manufacturer'] = lot.manufacturer
            # dict_lots['analytical_technique'] = lot.analytical_technique
            # dict_lots['reference_units'] = lot.reference_units
            # dict_lots['id_reactive'] = lot.id_reactive
            dict_lots['code_SAP'] = lot.code_SAP
            dict_lots['code_LOG'] = lot.code_LOG
            dict_lots['catalog_reference'] = lot.catalog_reference
            dict_lots['info_article'] = f"{lot.key}/-/{lot.catalog_reference}/-/{lot.manufacturer}/-/{lot.description}/-/{lot.analytical_technique}/-/{lot.reference_units}/-/{lot.id_reactive}/-/{lot.code_SAP}/-/{lot.code_LOG}/-/{lot.active}/-/{lot.temp_conservation}/-/{lot.description_subreference}/-/{lot.react_or_fungible}/-/{lot.code_panel}/-/{lot.location}/-/{lot.supplier}/-/{lot.purchase_format}/-/{lot.units_format}/-/{lot.import_unit_ics}/-/{lot.import_unit_idibgi}/-/{lot.local_management}/-/{lot.plataform_command_preferent}/-/{lot.maximum_amount}/-/{lot.purchase_format_supplier}/-/{lot.units_format_supplier}/-/{lot.name_logaritme}/-/{lot.units_for_discount}/-/{lot.units_measurement}/-/{lot.observations}/-/{lot.nif}/-/{lot.sales_contact}"
            dict_lots['description'] = lot.description
            dict_lots['description_subreference'] = lot.description_subreference
            dict_lots['active'] = lot.active
            # dict_lots['temp_conservation'] = lot.temp_conservation
            # dict_lots['react_or_fungible'] = lot.react_or_fungible

            list_lots.append(dict_lots)

        json_data = json.dumps(list_lots)
    except Exception:
        return "False_ No s'ha pogut accedir a la informació dels consums."

    return f'True_//_{json_data}'


@app.route('/history_lots', methods=['POST'])
@requires_auth
def history_lots():
    '''
        1 - Recollim la informació de l'ajax
        2 - Comprovem si aquest lot té història.
        2.1 - Si no en té retornem False més un missatge d'explicació per l'usuari.
        2.2 - Si és que si agafem la informació que hem trobat la posem en una llista de diccionaris.
        3 - Convertim la llista de diccionaris en un json
        4 - Retornem un True més la llista de diccionaris convertida a json.

        :param str historic_code_lot: Identificador unit del lot

        :return: True i la llista de diccionaris amb la info o False i una explicació per l'usuari
        :rtype: json
    '''
    historic_code_lot = request.form.get("historic_code_lot")

    try:
        info_history = session1.query(Stock_lots, Lot_consumptions).\
                       join(Lot_consumptions, Stock_lots.id == Lot_consumptions.id_lot).\
                       filter(Stock_lots.lot == historic_code_lot).\
                       all()
        if not info_history:
            return 'False_//_No hi ha informació sobre aquest lot.'

        list_consumptions = []
        for stock_lot, consumption in info_history:
            dict_consumption = {}
            dict_consumption['id'] = consumption.id
            dict_consumption['id_lot'] = consumption.id_lot
            if stock_lot.description_subreference == '':
                dict_consumption['description'] = stock_lot.description
            else:
                dict_consumption['description_subreference'] = stock_lot.description_subreference
            dict_consumption['lot'] = stock_lot.lot
            dict_consumption['catalog_reference'] = stock_lot.catalog_reference
            dict_consumption['internal_lot'] = stock_lot.internal_lot
            dict_consumption['date_open'] = consumption.date_open
            dict_consumption['user_open'] = consumption.user_open
            dict_consumption['date_close'] = consumption.date_close
            dict_consumption['user_close'] = consumption.user_close
            dict_consumption['observations_open'] = consumption.observations_open
            dict_consumption['observations_close'] = consumption.observations_close
            list_consumptions.append(dict_consumption)

        json_data = json.dumps(list_consumptions)
    except Exception:
        return "False_ No s'ha pogut accedir a la informació dels consums."

    return f'True_//_{json_data}'


@app.route('/search_fungible', methods=['POST'])
@requires_auth
def search_fungible():
    '''
        Realitza una cerca de lots fungibles basant-se en un codi introduït pel usuari.

        Aquesta funció utilitza el codi introduït per buscar lots fungibles en la base de dades. Primer cerca per la descripció 
        del lot, després cerca per la referència del catàleg, el codi SAP i el codi LOG si no es troben resultats inicials. 
        Si no es troben lots que coincideixin amb el codi, mostra un missatge d'advertència. Si ocorre un error durant la cerca, 
        mostra un missatge d'error.

        :param request: L'objecte de sol·licitud que conté el codi de cerca introduït pel usuari.
        :type request: flask.Request

        :return: Renderitza la plantilla `search_fungible.html` amb els lots seleccionats si la cerca té èxit, 
                o la plantilla `home.html` amb missatges d'advertència o error si no es troben lots o ocorre un error.
        :rtype: flask.Response
    '''
    code_search_fungible = request.form['code_search_fungible']

    try:
        if code_search_fungible == '':
            select_lots = session1.query(Stock_lots).filter_by(spent=0, react_or_fungible='Fungible').all()
        else:
            select_lots = session1.query(Stock_lots).filter_by(spent=0, react_or_fungible='Fungible', description=code_search_fungible).all()
            if not select_lots:
                select_lots = session1.query(Stock_lots).filter_by(spent=0, react_or_fungible='Fungible', catalog_reference=code_search_fungible).all()
                if not select_lots:
                    select_lots = session1.query(Stock_lots).filter_by(spent=0, react_or_fungible='Fungible', code_SAP=code_search_fungible).all()
                    if not select_lots:
                        select_lots = session1.query(Stock_lots).filter_by(spent=0, react_or_fungible='Fungible', code_LOG=code_search_fungible).all()

        if not select_lots:
            flash("No hi ha cap fungible amb el codi introduït", "warning")
            return render_template('home.html', list_desciption_lots=list_desciption_lots(),
                                   list_cost_center=list_cost_center())
    except Exception:
        flash("Error, no s'han pogut realitzar la cerca", "danger")
        return render_template('home.html', list_desciption_lots=list_desciption_lots(),
                               list_cost_center=list_cost_center())

    return render_template('search_fungible.html', select_lots=select_lots)


@app.route('/search_fungible_data', methods=['POST'])
@requires_auth
def search_fungible_data():
    code_search_fungible = request.form.get('code_search_fungible', '')

    try:
        if code_search_fungible == '':
            select_lots = session1.query(Stock_lots).filter_by(spent=0, react_or_fungible='Fungible').all()
        else:
            select_lots = session1.query(Stock_lots).filter_by(spent=0, react_or_fungible='Fungible', description=code_search_fungible).all()
            if not select_lots:
                select_lots = session1.query(Stock_lots).filter_by(spent=0, react_or_fungible='Fungible', catalog_reference=code_search_fungible).all()
                if not select_lots:
                    select_lots = session1.query(Stock_lots).filter_by(spent=0, react_or_fungible='Fungible', code_SAP=code_search_fungible).all()
                    if not select_lots:
                        select_lots = session1.query(Stock_lots).filter_by(spent=0, react_or_fungible='Fungible', code_LOG=code_search_fungible).all()

        if not select_lots:
            return jsonify({'success': False, 'message': "No hi ha cap fungible amb el codi introduït"})

        data = []
        for lot in select_lots:
            data.append({
                'id': lot.id,
                'catalog_reference': lot.catalog_reference or '',
                'id_reactive': lot.id_reactive or '',
                'code_SAP': lot.code_SAP or '',
                'code_LOG': lot.code_LOG or '',
                'description': lot.description_subreference or lot.description or '',
                'cost_center_stock': lot.cost_center_stock or '',
                'units_lot': lot.units_lot or '',
                'comand_number': lot.comand_number or '',
                'has_subreference': bool(lot.id_reactive)
            })
    except Exception:
        return jsonify({'success': False, 'message': "Error, no s'han pogut realitzar la cerca"})

    return jsonify({'success': True, 'data': data})


@app.route('/product_management_datalist')
@requires_auth
def product_management_datalist():
    rows = session1.query(Stock_lots).filter(Stock_lots.spent == 0).all()
    values = {'Reactiu': {}, 'Fungible': {}}

    for lot in rows:
        type_key = lot.react_or_fungible
        if type_key not in values:
            continue

        value = (lot.description or '').strip()
        if not value:
            continue

        if lot.id_reactive:
            display = f"{lot.catalog_reference or ''} - {lot.id_reactive} - {lot.description or ''}"
        else:
            display = f"{lot.catalog_reference or ''} - {lot.description or ''}"

        key = f"{value}__{display}"
        if key not in values[type_key]:
            values[type_key][key] = {'value': value, 'display': display}

    return jsonify({
        'success': True,
        'reactives': list(values['Reactiu'].values()),
        'fungibles': list(values['Fungible'].values())
    })


@app.route('/search_all_year', methods=['POST'])
@requires_auth
def search_all_year():
    '''
        Realitza una cerca de lots basant-se en un codi de cerca i la data de recepció de l'any actual.

        Aquesta funció busca lots en la base de dades utilitzant el codi de cerca proporcionat i l'any actual com a criteris.
        Primer busca per centre de cost, després per referència del catàleg, codi SAP, i finalment per la data de recepció
        si no es troben resultats. Si no es troben lots amb el codi introduït, retorna un missatge indicant que no s'ha trobat
        stock. Si ocorre un error durant la cerca, retorna un missatge d'error.

        :param request: L'objecte de sol·licitud que conté el codi de cerca introduït pel usuari.
        :type request: flask.Request

        :return: Una cadena amb el resultat de la cerca. Si la cerca és exitosa, retorna `True_//_{list_info_stock}` amb
                la informació del stock en format JSON. Si no es troben lots o ocorre un error, retorna un missatge d'error
                amb el prefix `False_//_`.
        :rtype: str
    '''
    search_data_code = request.form['search_data_code']

    # date = datetime.now()
    # year = date.strftime("-%Y")
    # list_year = [year, int(year)+1, int(year)+2, int(year)+3, int(year)+4]
    # print(year)
    # print(list_year)
    # try:
    if search_data_code == '':
        return 'False_//_Es codi no pot estar buit.'
    else:
        # select_lots = (
        #     session1.query(Stock_lots, Lot_consumptions)
        #     .outerjoin(Lot_consumptions, Lot_consumptions.id_lot == Stock_lots.id)
        #     .filter(
        #         or_(
        #             Stock_lots.cost_center_stock == search_data_code,
        #             Stock_lots.catalog_reference == search_data_code,
        #             Stock_lots.code_SAP == search_data_code,
        #             Stock_lots.reception_date == search_data_code.replace('/', '-')
        #         ),
        #         or_(*[Stock_lots.reception_date.like(f"%{year}%") for year in list_year])
        #     )
        #     .all()
        # )

        # select_lots = (
        #     session1.query(Stock_lots, Lot_consumptions)
        #     .outerjoin(Lot_consumptions, Lot_consumptions.id_lot == Stock_lots.id)
        #     .filter(
        #         or_(
        #             Stock_lots.cost_center_stock == search_data_code,
        #             Stock_lots.catalog_reference == search_data_code,
        #             Stock_lots.code_SAP == search_data_code,
        #             Stock_lots.reception_date == search_data_code.replace('/', '-')
        #         )
        #     ).all()
        # )

        if search_data_code != 'Tots':
            select_lots = (
                session1.query(Stock_lots, Lot_consumptions)
                .outerjoin(Lot_consumptions, Lot_consumptions.id_lot == Stock_lots.id)
                .filter(Stock_lots.reception_date.like(f'%{search_data_code}'))
                .all()
            )
        else:
            select_lots = (
                session1.query(Stock_lots, Lot_consumptions)
                .outerjoin(Lot_consumptions, Lot_consumptions.id_lot == Stock_lots.id)
                .all()
            )

    print(len(select_lots))
    if not select_lots:
        # return f"False_//_No s'ha trobat stock amb el codi {search_data_code}"
        return jsonify({"success": False, "data": f"No s'ha trobat stock de l'any {search_data_code}"})
    else:
        list_info_stock_aux = [
            {
                **to_dict(stock),
                **(to_dict(consumption) if consumption else {})  # si no hi ha consumption, afegeix dict buit
            } for stock, consumption in select_lots
        ]
        # list_info_stock = json.dumps(list_info_stock_aux)
    # except Exception:
    #     return "False_//_Error, no s'ha pout realitzar la cerca"

    # return f'True_//_{list_info_stock}'
    return jsonify({"success": True, "data": list_info_stock_aux})


@app.route('/download_certificate_pending', methods=['POST'])
@requires_auth
def download_certificate_pending():
    '''
    '''
    # try:
    # select_lot = session1.query(Stock_lots).filter(Stock_lots.react_or_fungible == 'Reactiu').group_by(Stock_lots.lot, Stock_lots.reception_date).all()
    # if not select_lot:
    #     flash("No s'ha trobat informació a la BD", "danger")
    #     return render_template('home.html', list_desciption_lots=list_desciption_lots(),
    #                             list_cost_center=list_cost_center())

    list_json = request.form['list_json']
                        
    # Crear un DataFrame con los datos
    data = {
        'Referencia Cataleg': [],
        'Descripció': [],
        'Codi subreferencia': [],
        'Descripció subref.': [],
        'Lot': [],
        'Lot intern': [],
        'Data recepció': [],
        'Data caducitat': [],
        'Observacions inspecció': [],
        'Proveidor': []
    }

    # for row in select_lot:
    #     if row.certificate != '':
    #         data['Referencia Cataleg'].append(row.catalog_reference)
    #         data['Descripció'].append(row.description)
    #         data['Codi subreferencia'].append(row.id_reactive)
    #         data['Descripció subref.'].append(row.description_subreference)
    #         data['Lot'].append(row.lot)
    #         data['Lot intern'].append(row.internal_lot)
    #         data['Data recepció'].append(str(row.reception_date))
    #         data['Data caducitat'].append(str(row.date_expiry))
    #         data['Observacions inspecció'].append(row.observations_inspection)

    list_data = json.loads(list_json)

    for row in list_data:
        data['Referencia Cataleg'].append(row['referencia_cataleg'])
        data['Descripció'].append(row['descripcio'])
        data['Codi subreferencia'].append(row['codi_subreferencia'])
        data['Descripció subref.'].append(row['descripcio_subref'])
        data['Lot'].append(row['lot'])
        data['Lot intern'].append(row['lot_intern'])
        data['Data recepció'].append(row['data_recepcio'])
        data['Data caducitat'].append(row['data_caducitat'])
        data['Observacions inspecció'].append(row['observacions_inspeccio'])
        data['Proveidor'].append(row['manufacturer'])

    df = pd.DataFrame(data)

    # Guardar el DataFrame en un archivo Excel
    path = f"{main_dir_docs}/plantillas/preus_articles.xlsx"
    df.to_excel(path, index=False)

    # Ajustar el tamaño de las columnas automáticamente
    wb = load_workbook(path)  # Cargar el archivo Excel
    ws = wb.active  # Obtener la hoja activa

    # --- Aplicar estilos al encabezado ---
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo amarillo
    header_font = Font(size=13, bold=True)  # Letra tamaño 13 y en negrita
    for cell in ws[1]:  # Primera fila (encabezado)
        cell.fill = header_fill
        cell.font = header_font

    # --- Ajustar el tamaño de las columnas automáticamente ---
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho

    ws.column_dimensions['B'].width = 24  # Ajustar ancho
    ws.column_dimensions['F'].width = 16  # Ajustar ancho
    ws.column_dimensions['G'].width = 17  # Ajustar ancho

    # --- Ajustar altura de todas las filas (margen superior e inferior) ---
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 19  # Ajustar altura de fila
        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="left")  # Alineación vertical centrada

    wb.save(path)  # Guardar cambios
    return send_file(path, as_attachment=True)
    # except Exception:
    #     flash("Error inesperat, contacteu amb un administrador", "danger")
    #     return render_template('home.html', list_desciption_lots=list_desciption_lots(),
    #                            list_cost_center=list_cost_center())


@app.route("/info_description_lots")
@requires_auth
def info_description_lots():
    lots = list_desciption_lots()

    data = [
        {
            "catalog_reference": x.catalog_reference,
            "description": x.description,
            "analytical_technique": x.analytical_technique,
            "id_reactive": x.id_reactive,
            "description_subreference": x.description_subreference,
            "code_panel": x.code_panel,
            "name_logaritme": x.name_logaritme,
            "supplier": x.supplier,
        }
        for x in lots
    ]
    return jsonify(data)


@app.route("/info_management_primers")
@requires_auth
def info_management_primers():
    lots = list_desciption_lots()
    primers = session1.query(Buy_primers).filter(Buy_primers.received == 0).filter(Buy_primers.delete == 0).all()

    data = [
        {
            "id": x.id,
            "dna": x.dna,
            "sequence_name": x.sequence_name,
            "purification": x.purification,
            "synthesis_scale": x.synthesis_scale,
            "shipping_conditions": x.shipping_conditions,
            "modification_5": x.modification_5,
            "sequence": x.sequence,
            "modification_3": x.modification_3,
            "internal_modification": x.internal_modification,
            "quality_check_maldi": x.quality_check_maldi,
            "technique": x.technique,
            "buy": x.buy,
            "buy_date": x.buy_date,
            "received": x.received,
            "received_date": x.received_date,
            "observations": x.observations,
        }
        for x in primers
    ]
    return jsonify(data)


@app.route('/add_buy_primer', methods=['POST'])
@requires_auth
def add_buy_primer():
    """
        Afegeix una nova compra de primer a la base de dades.

        :return: Resposta JSON amb l'estat de l'operació.
        :rtype: flask.Response
    """
    dna = request.form.get('dna', '')
    sequence_name = request.form.get('sequence_name', '')
    purification = request.form.get('purification', '')
    synthesis_scale = request.form.get('synthesis_scale', '')
    shipping_conditions = request.form.get('shipping_conditions', '')
    modification_5 = request.form.get('modification_5', '')
    sequence = request.form.get('sequence', '')
    modification_3 = request.form.get('modification_3', '')
    internal_modification = request.form.get('internal_modification', '')
    quality_check_maldi = request.form.get('quality_check_maldi', '')
    technique = request.form.get('technique', '')
    observations = request.form.get('observations', '')

    try:
        select_primer = session1.query(Buy_primers)\
            .filter(Buy_primers.sequence == sequence)\
            .filter(Buy_primers.received == 0)\
            .filter(Buy_primers.delete == 0)\
            .first()

        if select_primer:
            return jsonify({
                "status": "duplicate",
                "message": "Ja tenim un primer posat a comprar amb la mateix seqüència",
                "id": "none"
            }), 200

        insert_buy_primer = Buy_primers(
            dna=dna,
            sequence_name=sequence_name,
            purification=purification,
            synthesis_scale=synthesis_scale,
            shipping_conditions=shipping_conditions,
            modification_5=modification_5,
            sequence=sequence,
            modification_3=modification_3,
            internal_modification=internal_modification,
            quality_check_maldi=quality_check_maldi,
            technique=technique,
            observations=observations,
            request_by=session['acronim'],
            request_date=instant_date(),
            buy=0,
            buy_by='',
            buy_date='',
            received=0,
            received_by='',
            received_date='',
            email_send=session['email'],
            delete=0,
            delete_by='',
            delete_date='',
        )

        session1.add(insert_buy_primer)
        session1.commit()

        new_id = insert_buy_primer.id

        return jsonify({
            "status": "success",
            "message": "Primer posat a comprar correctament",
            "id": new_id
        }), 200

    except Exception as e:
        session1.rollback()
        return jsonify({
            "status": "error",
            "message": "Error, no s'ha pogut inserir el primer a compres.",
            "id": 'none'
        }), 500


@app.route('/action_primer', methods=['POST'])
@requires_auth
def action_primer():
    id_primer = request.form.get('id_primer', '')
    action = request.form.get('action', '')

    list_id_primer = id_primer.split(';')
    if action == 'tramited':
        text_email = '<p style="margin-bottom:10px;">Els següents primers han estat comprats :</p>'
        text_header_email = 'Compra de primers'

        select_command = session1.query(Buy_primers).filter(Buy_primers.command_id.like('4430954989-%')).all()
        max_num = 0
        for row in select_command:
            if row.command_id:
                try:
                    num = int(row.command_id.split('-')[-1])
                    if num > max_num:
                        max_num = num
                except:
                    continue

        next_num = max_num + 1
        new_command_id = f"4430954989-{next_num}"
    elif action == 'buy':
        text_email = '<p style="margin-bottom:10px;">Hem rebut els seguents primers :</p>'
        text_header_email = 'Recepcio de primers'
    else:
        text_email = ''
        text_header_email = ''

    date_now = instant_date()
    not_found = ''
    id_primers_change = ''
    send_mail = []
    isoformes = []
    external_primer_verify = None
    defer_reception_to_primers = False
    try:
        selected_primers = []
        if action == 'buy':
            for primer_id in list_id_primer:
                select_primer_pair = session1.query(Buy_primers).filter(Buy_primers.id == primer_id).first()
                if select_primer_pair is not None:
                    selected_primers.append(select_primer_pair)

            primer_pair = get_selected_primer_sense_pair(selected_primers)
            if primer_pair:
                isoforma = ''
                isoforma_error = ''
                try:
                    isoforma = get_isoforma_by_gene_exon(primer_pair['sequence_name'])
                except Exception as e:
                    isoforma_error = str(e)

                isoformes.append({
                    'sequence_name': f"{primer_pair['sense_name']} / {primer_pair['antisense_name']}",
                    'isoforma': isoforma,
                    'error': isoforma_error
                })
                external_primer_verify = {
                    'url': f'{ip_address}:5005/external/insert-primer/verify',
                    'gen': primer_pair['gen'],
                    'sense': primer_pair['sense'] or '',
                    'antisense': primer_pair['antisense'] or '',
                    'qpcr': primer_pair['qpcr'],
                    'm13': primer_pair['m13'],
                    'isoforma': isoforma or '',
                    'lots_primer_ids': id_primer,
                    'lots_receive_url': f'{ip_address}:5017/external/receive-primers'
                }
                defer_reception_to_primers = True

        for primer_id in list_id_primer:
            select_primer = session1.query(Buy_primers).filter(Buy_primers.id == primer_id).first()
            if select_primer is None:
                not_found += id_primer + ';'
            else:
                if action == 'tramited' and select_primer.buy == 0:
                    select_primer.buy = 1
                    select_primer.buy_date = date_now
                    select_primer.buy_by = session['acronim']
                    select_primer.command_id = new_command_id

                    text_email += f"&nbsp;&nbsp;&nbsp;• {select_primer.sequence_name} - {select_primer.sequence}<br>"
                    id_primers_change += primer_id + ';'
                    if select_primer.email_send not in send_mail:
                        send_mail.append(select_primer.email_send)
                elif action == 'buy' and select_primer.buy == 1:
                    if defer_reception_to_primers:
                        continue

                    select_primer.received = 1
                    select_primer.received_date = date_now
                    select_primer.received_by = session['acronim']

                    text_email += f"&nbsp;&nbsp;&nbsp;• {select_primer.sequence_name} - {select_primer.sequence}<br>"
                    id_primers_change += primer_id + ';'
                    if select_primer.email_send not in send_mail:
                        send_mail.append(select_primer.email_send)
                elif action == 'delete':
                    select_primer.delete = 1
                    select_primer.delete_date = date_now
                    select_primer.delete_by = session['acronim']
                    id_primers_change += primer_id + ';'
        
        session1.commit()

        if len(send_mail) > 0:
            result_emails = ','.join(send_mail)
            print(result_emails)
            send_mail_generic(result_emails, text_email, text_header_email)

        id_primers_change = id_primers_change.rstrip(';')
        not_found = not_found.rstrip(';')

        return jsonify({
            "status": "success",
            "message": "Primer posat a comprar correctament",
            "id": id_primers_change,
            "not_found": not_found,
            "isoformes": isoformes,
            "external_primer_verify": external_primer_verify,
            "deferred": defer_reception_to_primers
        }), 200

    except Exception as e:
        session1.rollback()
        return jsonify({
            "status": "error",
            "message": str(e) or "Error, no s'ha pogut procesar la petició solicitada",
            "id": 'none',
            "not_found": 'none'
        }), 500


@app.route('/external/receive-primers', methods=['POST'])
def external_receive_primers():
    id_primer = request.form.get('id_primer') or request.form.get('lots_primer_ids') or ''
    received_by = request.form.get('received_by') or request.form.get('user') or session.get('acronim') or 'Primers'
    list_id_primer = [primer_id for primer_id in id_primer.split(';') if primer_id]

    if not list_id_primer:
        return jsonify({
            "status": "error",
            "message": "No s'ha rebut cap ID de primer per recepcionar.",
            "id": "",
            "not_found": ""
        }), 400

    try:
        received_result = receive_primers_by_ids(list_id_primer, received_by)
        session1.commit()

        if len(received_result['send_mail']) > 0:
            result_emails = ','.join(received_result['send_mail'])
            send_mail_generic(result_emails, received_result['text_email'], received_result['text_header_email'])

        return jsonify({
            "status": "success",
            "message": "Primers recepcionats correctament.",
            "id": received_result['id_primers_change'],
            "not_found": received_result['not_found']
        }), 200

    except Exception as e:
        session1.rollback()
        return jsonify({
            "status": "error",
            "message": str(e) or "No s'han pogut recepcionar els primers.",
            "id": "",
            "not_found": ""
        }), 500


@app.route("/upadate_bd")
@requires_auth
def upadate_bd():
    """
        Llegeix un fitxer Excel des del directori principal i processa les files per actualitzar la BD.

        El fitxer Excel s'espera que tingui capçaleres (noms de columna) a la primera fila.
        Es recorre fila a fila i es crea un diccionari amb tots els camps per a cada fila.

        :return: Resposta JSON amb el resultat del procés.
        :rtype: flask.wrappers.Response

        :raises ValueError: Si no es troba el fitxer o si falten columnes esperades.
    """
    modify = 0
    not_found = ''
    found = 0
    # --- Config ---
    # Comentari en català: ruta del fitxer Excel dins el directori principal del projecte
    csv_filename = "new_db.xlsx"  # <-- canvia-ho pel nom real
    csv_path = os.path.join(main_dir, csv_filename)

    if not os.path.exists(csv_path):
        return jsonify({"result": False, "message": f"No s'ha trobat l'Excel a: {csv_path}"}), 404

    print(csv_path)

    # try:
    df = pd.read_excel(csv_path)
    # except Exception as exc:
    #     return "Error al lleguir el document"

    # Comentari en català: neteja bàsica (files buides i NaN)
    df = df.dropna(how="all").fillna("")

    # (Opcional però recomanat) validar que tens les 11 columnes esperades
    # expected_cols = ["ID Petició", "...", "..."]  # posa aquí les 11 columnes reals
    # missing = [c for c in expected_cols if c not in df.columns]
    # if missing:
    #     return jsonify({"result": False, "message": f"Falten columnes: {missing}"}), 400

    not_match: list[str] = []
    updated: int = 0

    # Comentari en català: iteració fila a fila; cada fila és un dict amb totes les columnes
    for row_dict in df.to_dict(orient="records"):
        dict_log = {}
        # Exemple: obtenir un camp concret
        key_raw = row_dict.get("key", "")
        key = str(key_raw).strip()

        catalog_reference_raw = row_dict.get("catalog_reference", "")
        catalog_reference = str(catalog_reference_raw).strip()

        manufacturer_raw = row_dict.get("manufacturer", "")
        manufacturer = str(manufacturer_raw).strip()

        description_raw = row_dict.get("description", "")
        description = str(description_raw).strip()

        analytical_technique_raw = row_dict.get("analytical_technique", "")
        analytical_technique = str(analytical_technique_raw).strip()

        reference_units_raw = row_dict.get("reference_units", "")
        reference_units = str(reference_units_raw).strip()

        id_reactive_raw = row_dict.get("id_reactive", "")
        id_reactive = str(id_reactive_raw).strip()

        code_SAP_raw = row_dict.get("code_SAP", "")
        code_SAP = str(code_SAP_raw).strip()

        code_LOG_raw = row_dict.get("code_LOG", "")
        code_LOG = str(code_LOG_raw).strip()

        active_raw = row_dict.get("active", "")
        active = str(active_raw).strip()

        temp_conservation_raw = row_dict.get("temp_conservation", "")
        temp_conservation = str(temp_conservation_raw).strip()

        description_subreference_raw = row_dict.get("description_subreference", "")
        description_subreference = str(description_subreference_raw).strip()

        react_or_fungible_raw = row_dict.get("react_or_fungible", "")
        react_or_fungible = str(react_or_fungible_raw).strip()

        code_panel_raw = row_dict.get("code_panel", "")
        code_panel = str(code_panel_raw).strip()

        location_raw = row_dict.get("location", "")
        location = str(location_raw).strip()

        supplier_raw = row_dict.get("supplier", "")
        supplier = str(supplier_raw).strip()

        purchase_format_raw = row_dict.get("purchase_format", "")
        purchase_format = str(purchase_format_raw).strip()

        units_format_raw = row_dict.get("units_format", "")
        units_format = str(units_format_raw).strip()

        import_unit_ics_raw = row_dict.get("import_unit_ics", "")
        import_unit_ics = str(import_unit_ics_raw).strip()

        import_unit_idibgi_raw = row_dict.get("import_unit_idibgi", "")
        import_unit_idibgi = str(import_unit_idibgi_raw).strip()

        local_management_raw = row_dict.get("local_management", "")
        local_management = str(local_management_raw).strip()

        plataform_command_preferent_raw = row_dict.get("plataform_command_preferent", "")
        plataform_command_preferent = str(plataform_command_preferent_raw).strip()

        maximum_amount_raw = row_dict.get("maximum_amount", "")
        maximum_amount = str(maximum_amount_raw).strip()

        purchase_format_supplier_raw = row_dict.get("purchase_format_supplier", "")
        purchase_format_supplier = str(purchase_format_supplier_raw).strip()

        units_format_supplier_raw = row_dict.get("units_format_supplier", "")
        units_format_supplier = str(units_format_supplier_raw).strip()

        nom_logaritme_raw = row_dict.get("Nom logaritme", "")
        name_logaritme = str(nom_logaritme_raw).strip()

        ubicació_raw = row_dict.get("Ubicació", "")
        ubicació = str(ubicació_raw).strip()

        unitats_raw = row_dict.get("Unitats", "")
        units = str(unitats_raw).strip()

        unitats_de_mesuta_raw = row_dict.get("Unitats de Mesuta", "")
        units_measurement = str(unitats_de_mesuta_raw).strip()

        observacions_raw = row_dict.get("Observacions", "")
        observations = str(observacions_raw).strip()

        select_lot = session1.query(Lots).filter(Lots.key == key).first()
        if not select_lot:
            not_found += f"<br>{key}"
        else:
            found += 1
            def update_lot_field(field, new_value, log_value=None):
                old_value = getattr(select_lot, field)
                if old_value != new_value:
                    setattr(select_lot, field, new_value)
                    dict_log[f'{field}_old'] = old_value
                    dict_log[f'{field}_new'] = log_value if log_value is not None else new_value

            update_lot_field('catalog_reference', catalog_reference)
            update_lot_field('manufacturer', manufacturer)
            update_lot_field('description', description)
            update_lot_field('analytical_technique', analytical_technique)
            update_lot_field('reference_units', reference_units)
            update_lot_field('id_reactive', id_reactive)
            update_lot_field('code_SAP', code_SAP)
            update_lot_field('code_LOG', code_LOG)
            update_lot_field('active', int(active), active)
            update_lot_field('temp_conservation', temp_conservation)
            update_lot_field('description_subreference', description_subreference)
            update_lot_field('react_or_fungible', react_or_fungible)
            update_lot_field('code_panel', code_panel)
            update_lot_field('location', location)
            update_lot_field('supplier', supplier)
            update_lot_field('purchase_format', purchase_format)
            update_lot_field('units_format', int(units_format), units_format)
            update_lot_field('import_unit_ics', import_unit_ics)
            update_lot_field('import_unit_idibgi', import_unit_idibgi)
            update_lot_field('local_management', local_management)
            update_lot_field('plataform_command_preferent', plataform_command_preferent)
            try:
                update_lot_field('maximum_amount', int(maximum_amount), maximum_amount)
            except:
                print(select_lot.maximum_amount)
            update_lot_field('purchase_format_supplier', purchase_format_supplier)
            try:
                update_lot_field('units_format_supplier', int(units_format_supplier), units_format_supplier)
            except:
                print(select_lot.units_format_supplier)

            # if select_lot.name_logaritme != name_logaritme:
            select_lot.name_logaritme = name_logaritme
            #     dict_log['name_logaritme_new'] = name_logaritme
            #     dict_log['name_logaritme_old'] = select_lot.name_logaritme

            # if select_lot.units_for_discount != units:
            select_lot.units_for_discount = units
            #     dict_log['units_new'] = units
            #     dict_log['units_old'] = select_lot.units

            # if select_lot.units_measurement != units_measurement:
            select_lot.units_measurement = units_measurement
            #     dict_log['units_measurement_new'] = units_measurement
            #     dict_log['units_measurement_old'] = select_lot.units_measurement

            # if select_lot.observations != observations:
            select_lot.observations = observations
            #     dict_log['observations_new'] = observations
            #     dict_log['observations_old'] = select_lot.observations

            if ubicació != '':
                update_lot_field('location', ubicació)

            select_stock = session1.query(Stock_lots).filter(Stock_lots.id_lot == key).all()
            for stock in select_stock:
                stock.catalog_reference = catalog_reference
                stock.manufacturer = manufacturer
                stock.description = description
                stock.analytical_technique = analytical_technique
                stock.reference_units = reference_units
                stock.id_reactive = id_reactive
                stock.code_SAP = code_SAP
                stock.code_LOG = code_LOG
                stock.active = active
                stock.temp_conservation = temp_conservation
                stock.description_subreference = description_subreference
                stock.react_or_fungible = react_or_fungible
                stock.code_panel = code_panel
                stock.location = location
                stock.supplier = supplier
                stock.purchase_format = purchase_format
                stock.units_format = units_format
                stock.local_management = local_management
                stock.plataform_command_preferent = plataform_command_preferent
                stock.maximum_amount = maximum_amount
                stock.purchase_format_supplier = purchase_format_supplier
                stock.units_format_supplier = units_format_supplier
                stock.name_logaritme = name_logaritme
                stock.units_for_discount = units
                stock.units_measurement = units_measurement
                stock.observations = observations
                
                if ubicació != '':
                    stock.location = ubicació

            if dict_log:
                modify += 1
                dict_save_info = {'id_lot': key,
                                  'type': 'update_bd_excel',
                                  'user': session['acronim'],
                                  'info': json.dumps(dict_log),
                                  'id_user': session['idClient'],
                                  'date': instant_date()}

                save_log(dict_save_info)

        session1.commit()

    return f"True<br>trobats ->{found}<br>Modificats ->{modify}<br>No trobats -> {not_found}"


@app.route("/info_commands_primers")
@requires_auth
def info_commands_primers():
    list_command_id = []
    select_command = session1.query(Buy_primers).filter(Buy_primers.command_id.like('4430954989-%')).all()
    for command_primer in select_command:
        if command_primer.command_id not in list_command_id:
            list_command_id.append(command_primer.command_id)

    list_command_id = sorted(
        {row.command_id for row in select_command if row.command_id},
        key=lambda x: int(x.split('-')[-1]),
        reverse=True
    )

    return jsonify(list_command_id)


def create_openpyxl_safe_template_copy(template_path):
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()

    def remove_invalid_merge_cell(match):
        ref_match = re.search(rb'\bref="([^"]+)"', match.group(0))
        if not ref_match:
            return b''

        ref = ref_match.group(1)
        if re.match(rb'^[A-Z]{1,3}[0-9]+:[A-Z]{1,3}[0-9]+$', ref):
            return match.group(0)

        return b''

    with zipfile.ZipFile(template_path, 'r') as source_zip, zipfile.ZipFile(temp_file.name, 'w', zipfile.ZIP_DEFLATED) as target_zip:
        for item in source_zip.infolist():
            content = source_zip.read(item.filename)

            if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                content = re.sub(rb'<mergeCell[^>]*/>', remove_invalid_merge_cell, content)
                content = re.sub(rb'<mergeCell[^>]*></mergeCell>', remove_invalid_merge_cell, content)
                content = re.sub(rb'<mergeCells[^>]*>\s*</mergeCells>', b'', content)

            target_zip.writestr(item, content)

    return temp_file.name


@app.route('/create_excel_primer', methods=['POST'])
@requires_auth
def create_excel_primer():
    """
        Crea i descarrega un arxiu Excel amb etiquetes d'ADN o batch.

        :return: Fitxer Excel o missatge d'error.
        :rtype: Response
    """
    try:
        command_id_primer = request.form.get('command_id_primer')
        command_primers_selected = request.form.get('command_primers_selected')
        if command_primers_selected == '' and not command_id_primer:
            return jsonify({
                'success': False,
                'message': "No s'ha pogut crear l'Excel: no s'ha rebut cap primer seleccionat ni cap comanda."
            }), 400

        if command_primers_selected != '':
            select_primer = []
            list_primers_selected = command_primers_selected.split(';')
            for primers_sel in list_primers_selected:
                print(primers_sel)
                select_primer_sel = session1.query(Buy_primers).filter(Buy_primers.id == primers_sel).first()
                if select_primer_sel is not None:
                    select_primer.append(select_primer_sel)
                else:
                    return jsonify({
                        'success': False,
                        'message': f"No s'ha pogut crear l'Excel: el primer amb ID {primers_sel} no existeix o no es pot recuperar."
                    }), 404
        else:
            select_primer = session1.query(Buy_primers).filter(Buy_primers.command_id == command_id_primer).all()

        if not select_primer:
            return jsonify({
                'success': False,
                'message': "No s'ha pogut crear l'Excel: no s'ha trobat cap primer amb les dades indicades."
            }), 404

        template_path = f"{main_dir_docs}/plantillas/plantilla_primers.xlsx"
        if not os.path.exists(template_path):
            return jsonify({
                'success': False,
                'message': f"No s'ha pogut crear l'Excel: no s'ha trobat la plantilla plantilla_primers.xlsx a {template_path}."
            }), 404

        safe_template_path = create_openpyxl_safe_template_copy(template_path)
        try:
            wb = load_workbook(safe_template_path)
        finally:
            if os.path.exists(safe_template_path):
                os.remove(safe_template_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=8):
            for cell in row:
                cell.value = None

        start = 2
        for primer in select_primer:
            sheet.cell(row=start, column=1, value=primer.sequence_name)
            sheet.cell(row=start, column=2, value=primer.synthesis_scale)
            sheet.cell(row=start, column=3, value=primer.shipping_conditions)
            sheet.cell(row=start, column=4, value=primer.purification)
            sheet.cell(row=start, column=5, value=primer.modification_5)
            sheet.cell(row=start, column=6, value=primer.sequence)
            sheet.cell(row=start, column=7, value=primer.modification_3)
            sheet.cell(row=start, column=8, value=primer.internal_modification)
            start += 1

        name_doc_buy_primers = 'plantilla_primers'
        output_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        output_file.close()
        wb.save(output_file.name)

        @after_this_request
        def cleanup_created_primer_excel(response):
            try:
                os.remove(output_file.name)
            except OSError:
                pass
            return response

        return send_file(output_file.name, as_attachment=True, download_name=f"{name_doc_buy_primers}.xlsx")
    except zipfile.BadZipFile:
        return jsonify({
            'success': False,
            'message': "No s'ha pogut crear l'Excel: la plantilla plantilla_primers.xlsx no és un fitxer Excel vàlid."
        }), 500
    except PermissionError:
        return jsonify({
            'success': False,
            'message': "No s'ha pogut crear l'Excel: no hi ha permisos per llegir la plantilla o guardar el fitxer."
        }), 500
    except OSError as error:
        return jsonify({
            'success': False,
            'message': f"No s'ha pogut crear l'Excel: error del sistema de fitxers ({error})."
        }), 500
    except Exception as error:
        return jsonify({
            'success': False,
            'message': f"No s'ha pogut crear l'Excel de primers: {error}"
        }), 500


@app.route('/get_lots', methods=['GET'])
def get_lots():
    """
    Retorna una llista de lots.

    :return: JSON amb la llista de lots.
    :rtype: flask.Response
    """
    lots = session1.query(Lots).group_by(Lots.manufacturer).order_by(func.lower(Lots.manufacturer)).all()

    result = []
    for lot in lots:
        result.append({
            "manufacturer": lot.manufacturer,
            "nif": lot.nif,
            "sales_contact": lot.sales_contact
        })

    return jsonify(result)


'''@app.route('/charge_excel')
def charge_excel():
    try:
        # Llegim el directori i el convertim en una llista
        directori = f'{main_dir}/info.xlsx'
        print(directori)
        df = pd.read_excel(directori, header=None)
        list_excel = df.values.tolist()
    except Exception:
        print("No s'ha pogut lleguir el document")
        return False
    print(f'linia maxima excel --> {len(list_excel)}')

    for line in range(1, len(list_excel)):
        print(list_excel[line])
        try:
            insert_lots = Lots(catalog_reference=str(list_excel[line][6]) if str(list_excel[line][6]) != 'nan' else '',
                               manufacturer=str(list_excel[line][3]) if str(list_excel[line][3]) != 'nan' else '',
                               description=str(list_excel[line][5]) if str(list_excel[line][5]) != 'nan' else '',
                            analytical_technique=str(list_excel[line][0]) if str(list_excel[line][0]) != 'nan' else '',
                               reference_units=1,
                               id_reactive='',
                               code_SAP=str(list_excel[line][8]) if str(list_excel[line][8]) != 'nan' else '',
                               code_LOG=str(list_excel[line][7]) if str(list_excel[line][7]) != 'nan' else '',
                               active=1,
                               temp_conservation=str(list_excel[line][9]) if str(list_excel[line][9]) != 'nan' else '',
                               description_subreference='',
                               react_or_fungible=str(list_excel[line][2]) if str(list_excel[line][2]) != 'nan' else '',
                               code_panel='')
            session1.add(insert_lots)
        except Exception:
            print("error")
    session1.commit()
    return "fet" '''


'''@app.route('/add_samples_for_excel')
@requires_auth
def add_samples_for_excel():
    # directori = f'{main_dir}/doc_nuria_articles_unics.xlsx'
    # directori = f'{main_dir}/doc_nuria_all.xlsx'
    # df = pd.read_excel(directori)
    # df = pd.read_excel(directori, header=None)
    # list_excel = df.values.tolist()
    # print("inici del documenttttt")
    # list = []
    # list_dupl = []
    # for line in range(1, len(list_excel)):
    #     if str(list_excel[line][2]).rstrip() not in list:
    #             list.append(str(list_excel[line][2]))
        # if list_excel[line][11] == 'SI':
        #     # Esto ha sido para seleccionar los que tiene subreferencias i hacer-les un update de los datos conjuntos
        #     # select_lot = session1.query(Lots).filter(Lots.catalog_reference == str(list_excel[line][2]).rstrip()).all()
        #     # if select_lot:

        #     select_lot = session1.query(Lots).filter(Lots.catalog_reference == str(list_excel[line][2]).rstrip()).first()
        #     if select_lot is not None:
        #         # print(f"El lot {list_excel[line][2]} ja esta introduit, -> {select_lot.key}")
        #         for select in select_lot:
        #             select.manufacturer = list_excel[line][23]
        #             select.description = list_excel[line][1]
        #             select.analytical_technique = list_excel[line][0]
        #             select.reference_units = 1
        #             select.id_reactive = ''
        #             select.code_SAP = list_excel[line][4]
        #             select.code_LOG = list_excel[line][3]
        #             select.active = 1
        #             select.temp_conservation = list_excel[line][13]
        #             select.description_subreference = ''
        #             select.react_or_fungible = list_excel[line][10]
        #             select.location = list_excel[line][14]
        #             select.supplier = list_excel[line][20]
        #             select.purchase_format = list_excel[line][9]
        #             select.units_format = list_excel[line][8]
        #             select.import_unit_ics = list_excel[line][21]
        #             select.import_unit_idibgi = list_excel[line][22]
        #             select.local_management = list_excel[line][7]
        #             select.plataform_command_preferent = list_excel[line][5]
        #             select.maximum_amount = list_excel[line][19]
        #             select.purchase_format_supplier = list_excel[line][16]
        #             select.units_format_supplier = list_excel[line][15]
        #         session1.commit()
        #     else:
        #         print("ha entrat a noussssssssssssssssssssssssssssssss")
        #         for i in range(list_excel[line][12]):
        #             insert_lot = Lots(catalog_reference=str(list_excel[line][2]).rstrip(),
        #                               manufacturer=list_excel[line][23],
        #                               description=list_excel[line][1],
        #                               analytical_technique=list_excel[line][0],
        #                               reference_units=1,
        #                               id_reactive='',
        #                               code_SAP=list_excel[line][4],
        #                               code_LOG=list_excel[line][3],
        #                               active=1,
        #                               temp_conservation=list_excel[line][13],
        #                               description_subreference='',
        #                               react_or_fungible=list_excel[line][10],
        #                               code_panel='',
        #                               location=list_excel[line][14],
        #                               supplier=list_excel[line][20],
        #                               purchase_format=list_excel[line][9],
        #                               units_format=list_excel[line][8],
        #                               import_unit_ics=list_excel[line][21],
        #                               import_unit_idibgi=list_excel[line][22],
        #                               local_management=list_excel[line][7],
        #                               plataform_command_preferent=list_excel[line][5],
        #                               maximum_amount=list_excel[line][19],
        #                               purchase_format_supplier=list_excel[line][16],
        #                               units_format_supplier=list_excel[line][15]
        #             )

        #             session1.add(insert_lot)
        #         session1.commit()
        # else:
        #     # aixo es per saber quins tenia duplicats a l'excel
        #     # if str(list_excel[line][2]).rstrip() not in list:
        #     #     list.append(str(list_excel[line][2]).rstrip())
        #     # else:
        #     #     if str(list_excel[line][2]).rstrip() not in list_dupl:
        #     #         list_dupl.append(str(list_excel[line][2]).rstrip())

        #     # Aqui inserirem o actualitzarem la info
        #     select_lot = session1.query(Lots).filter(Lots.catalog_reference == str(list_excel[line][2]).rstrip()).all()
        #     if select_lot:
        #         if len(select_lot) > 1:
        #             print(f"El lot {list_excel[line][2]} se nan trobat 2")
        #         else:
        #             for select in select_lot:
        #                 select.manufacturer = list_excel[line][23]
        #                 select.description = list_excel[line][1]
        #                 select.analytical_technique = list_excel[line][0]
        #                 select.reference_units = 1
        #                 select.id_reactive = ''
        #                 select.code_SAP = list_excel[line][4]
        #                 select.code_LOG = list_excel[line][3]
        #                 select.active = 1
        #                 select.temp_conservation = list_excel[line][13]
        #                 select.description_subreference = ''
        #                 select.react_or_fungible = list_excel[line][10]
        #                 select.location = list_excel[line][14]
        #                 select.supplier = list_excel[line][20]
        #                 select.purchase_format = list_excel[line][9]
        #                 select.units_format = list_excel[line][8]
        #                 select.import_unit_ics = list_excel[line][21]
        #                 select.import_unit_idibgi = list_excel[line][22]
        #                 select.local_management = list_excel[line][7]
        #                 select.plataform_command_preferent = list_excel[line][5]
        #                 select.maximum_amount = list_excel[line][19]
        #                 select.purchase_format_supplier = list_excel[line][16]
        #                 select.units_format_supplier = list_excel[line][15]
        #             session1.commit()
        #     else:
        #         print("ha entrat a noussssssssssssssssssssssssssssssss")
        #         insert_lot = Lots(catalog_reference=str(list_excel[line][2]).rstrip(),
        #                           manufacturer=list_excel[line][23],
        #                           description=list_excel[line][1],
        #                           analytical_technique=list_excel[line][0],
        #                           reference_units=1,
        #                           id_reactive='',
        #                           code_SAP=list_excel[line][4],
        #                           code_LOG=list_excel[line][3],
        #                           active=1,
        #                           temp_conservation=list_excel[line][13],
        #                           description_subreference='',
        #                           react_or_fungible=list_excel[line][10],
        #                           code_panel='',
        #                           location=list_excel[line][14],
        #                           supplier=list_excel[line][20],
        #                           purchase_format=list_excel[line][9],
        #                           units_format=list_excel[line][8],
        #                           import_unit_ics=list_excel[line][21],
        #                           import_unit_idibgi=list_excel[line][22],
        #                           local_management=list_excel[line][7],
        #                           plataform_command_preferent=list_excel[line][5],
        #                           maximum_amount=list_excel[line][19],
        #                           purchase_format_supplier=list_excel[line][16],
        #                           units_format_supplier=list_excel[line][15])
        #         session1.add(insert_lot)
        #         session1.commit()

    # select_lot = session1.query(Lots).all()
    # for lot in select_lot:
    #     if lot.catalog_reference not in list:
    #         if lot.catalog_reference not in list_dupl:
    #             list_dupl.append(lot.catalog_reference)

    select = session1.query(Stock_lots).all()s

    return "fet" '''
